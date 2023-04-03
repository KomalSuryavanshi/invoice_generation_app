import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path

filepaths = glob.glob("invoices/*.xlsx")  # Use glob to read all xlsx files

for filepath in filepaths:

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()

    filename = Path(filepath).stem
    invoice_no, date = filename.split("-")   # filename.split("-") will give out ['10001','2023.1.18'] so add [0] to get just the 10001

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, txt=f"Invoice No.{invoice_no}", ln=1)

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, txt=f"Date :- {date}", ln=1)

    df = pd.read_excel(filepath, sheet_name="Sheet 1")
    # Add a header
    title_of_cell = df.columns
    title_of_cell = [item.replace("_", " ").title() for item in title_of_cell]
    pdf.set_font(family="Times", size=10,style="B")
    pdf.set_text_color(80, 80, 80)

    pdf.cell(w=30, h=8, txt=title_of_cell[0], border=1)
    pdf.cell(w=40, h=8, txt=title_of_cell[1], border=1)
    pdf.cell(w=60, h=8, txt=title_of_cell[2], border=1)
    pdf.cell(w=40, h=8, txt=title_of_cell[3], border=1)
    pdf.cell(w=40, h=8, txt=title_of_cell[4], border=1, ln=1)

    # Adding elements in the cell
    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80, 80, 80)

        pdf.cell(w=30, h=10, txt=str(row["product_id"]), border=1)
        pdf.cell(w=40, h=10, txt=str(row["product_name"]), border=1)
        pdf.cell(w=60, h=10, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=40, h=10, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=40, h=10, txt=str(row["total_price"]), border=1, ln=1)

    #  Adding sum of total products
    total_sum = df["total_price"].sum()

    pdf.set_font(family="Times", size=10, style="B")
    pdf.set_text_color(80, 80, 80)
    # To set other columns empty and only add total at last column
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=40, h=8, txt="", border=1)
    pdf.cell(w=60, h=8, txt="", border=1)
    pdf.cell(w=40, h=8, txt="Total ", border=1)
    pdf.cell(w=40, h=8, txt=str(total_sum), border=1, ln=1)

    # Adding total sum sentence
    pdf.set_font(family="Times", size=12, style="B")
    pdf.set_text_color(100, 100, 100)
    pdf.cell(w=30, h=8, txt=f"The total price is : {total_sum}", ln=1)

    pdf.output(f"PDFs/{filename}.pdf")
